"""Bounded XLSX workbook parsing for governed dataset ingestion.

Uses openpyxl in read-only, data_only mode so formula cells yield cached values
and workbook macros/code are never executed. Imports the first worksheet unless
an explicit sheet name is provided.
"""

from __future__ import annotations

import re
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from vip_api.core.errors import ApplicationError

# Unicode-aware identifiers for quoted PostgreSQL columns (Arabic included).
HEADER = re.compile(r"^[^\W\d]\w{0,62}$", re.UNICODE)
_MAX_ROWS = 50_000
_MAX_COLS = 200


def _cell_to_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, datetime):
        if value.time() == time(0, 0):
            return value.date().isoformat()
        return value.replace(microsecond=0).isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.replace(microsecond=0).isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, float):
        if value != value:  # NaN
            return None
        if value.is_integer() and abs(value) < 1e15:
            return str(int(value))
        return format(value, ".15g")
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    return text if text != "" else None


def _normalize_headers(raw: list[str | None]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(raw):
        candidate = (value or "").strip()
        if not candidate or not HEADER.fullmatch(candidate):
            candidate = f"column_{index + 1}"
        base = candidate[:63]
        count = seen.get(base, 0)
        seen[base] = count + 1
        if count:
            suffix = f"_{count + 1}"
            candidate = f"{base[: 63 - len(suffix)]}{suffix}"
        else:
            candidate = base
        headers.append(candidate)
    if not headers or len(headers) > _MAX_COLS or len(set(headers)) != len(headers):
        raise ApplicationError(
            code="XLSX_HEADERS_INVALID",
            message="Workbook headers must be unique and contain at most 200 columns.",
            status_code=422,
        )
    return headers


def parse_xlsx(
    path: Path,
    *,
    sheet_name: str | None = None,
) -> tuple[list[str], list[list[str | None]], str]:
    """Return (headers, rows, sheet_used).

    Formula policy: data_only=True returns cached calculated values. Cells with
    no cached result become null. Macros are never executed.
    """
    if path.stat().st_size == 0:
        raise ApplicationError(
            code="XLSX_EMPTY", message="The Excel workbook is empty.", status_code=422
        )
    try:
        workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    except InvalidFileException as exc:
        raise ApplicationError(
            code="XLSX_INVALID",
            message="The Excel workbook could not be opened.",
            status_code=422,
        ) from exc
    except Exception as exc:
        message = str(exc).lower()
        if "password" in message or "encrypted" in message:
            raise ApplicationError(
                code="XLSX_ENCRYPTED",
                message="Password-protected Excel workbooks are not supported.",
                status_code=422,
            ) from exc
        raise ApplicationError(
            code="XLSX_INVALID",
            message="The Excel workbook could not be opened.",
            status_code=422,
        ) from exc

    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ApplicationError(
                    code="XLSX_SHEET_NOT_FOUND",
                    message="The requested worksheet was not found.",
                    status_code=422,
                )
            sheet = workbook[sheet_name]
            used = sheet_name
        else:
            if not workbook.sheetnames:
                raise ApplicationError(
                    code="XLSX_EMPTY",
                    message="The Excel workbook has no worksheets.",
                    status_code=422,
                )
            used = workbook.sheetnames[0]
            sheet = workbook[used]

        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration as exc:
            raise ApplicationError(
                code="XLSX_EMPTY",
                message="The Excel worksheet must include a header row.",
                status_code=422,
            ) from exc

        # Trim trailing empty header cells.
        header_values = list(header_row)
        while header_values and header_values[-1] is None:
            header_values.pop()
        if not header_values:
            raise ApplicationError(
                code="XLSX_HEADERS_INVALID",
                message="The Excel worksheet header row is empty.",
                status_code=422,
            )
        headers = _normalize_headers([_cell_to_text(cell) for cell in header_values])
        width = len(headers)

        rows: list[list[str | None]] = []
        for number, row in enumerate(rows_iter, start=2):
            if number > _MAX_ROWS + 1:
                raise ApplicationError(
                    code="XLSX_ROW_LIMIT",
                    message="Interactive XLSX ingestion is limited to 50,000 rows.",
                    status_code=422,
                )
            values = [_cell_to_text(cell) for cell in row[:width]]
            if len(values) < width:
                values.extend([None] * (width - len(values)))
            if all(value is None for value in values):
                continue  # drop trailing / blank rows
            rows.append(values)

        # Drop trailing fully-empty rows already skipped; ensure at least one data row.
        if not rows:
            raise ApplicationError(
                code="XLSX_EMPTY",
                message="The Excel worksheet must include at least one data row.",
                status_code=422,
            )
        return headers, rows, used
    finally:
        workbook.close()
